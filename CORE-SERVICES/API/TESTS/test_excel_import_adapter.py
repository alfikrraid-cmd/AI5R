"""
MWO-LTSA-090 -- ExcelImportAdapter (API.import_adapter). Real .xlsx fixtures
throughout (written with the real, already-installed openpyxl, the same
library ExcelReader itself uses to read .xlsx). The one exception is the
"xls delegates correctly" test: no .xls-writing library (xlwt/pyexcel/etc.)
is installed in this environment to author a genuine legacy-BIFF fixture,
so that one test stubs only the raw `ExcelReader._read_xls` acquisition
call -- ExcelReader.read()'s own real suffix dispatch and
ExcelImportAdapter's real detect/map/build logic still run unmodified, end
to end, against a real .xls-suffixed file. The "corrupt .xls" test needs no
such stub: genuinely unreadable bytes make the real, unpatched
xlrd.open_workbook() fail on its own.
"""

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

CORE_SERVICES_DIR = Path(__file__).resolve().parents[2]
if str(CORE_SERVICES_DIR) not in sys.path:
    sys.path.insert(0, str(CORE_SERVICES_DIR))

from API.import_adapter import parse_import_file  # noqa: E402
from API.import_validator import ImportPackage  # noqa: E402
from ENTERPRISE_DATA_ENGINE import ExcelReader  # noqa: E402
from FACTORY.CORE.exceptions import ManufacturingValidationError  # noqa: E402

_EMPTY_PACKAGE = ImportPackage(pumps=(), seals=(), installations=(), documents=())


def _save_workbook(path: Path, sheets):
    """sheets: sequence of (name, rows). Same shape as ENTERPRISE_DATA_ENGINE/
    TESTS/test_excel_reader.py's own save_workbook() helper, reused at the
    same shape rather than re-derived."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        worksheet = workbook.create_sheet(name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)


# --- xlsx / xls delegation --------------------------------------------------------


def test_xlsx_delegates_through_the_real_excel_reader(tmp_path):
    path = tmp_path / "package.xlsx"
    _save_workbook(path, [("Pump", (("tag_number", "area"), ("P-100", "Unit 1")))])

    result = parse_import_file(path)

    assert isinstance(result, ImportPackage)
    assert result.pumps == ({"tag_number": "P-100", "area": "Unit 1"},)


def test_xls_delegates_correctly(tmp_path, monkeypatch):
    path = tmp_path / "package.xls"
    path.write_bytes(b"placeholder")  # content irrelevant -- _read_xls is stubbed below

    def fake_read_xls(_path):
        return (
            {"name": "Seal", "rows": (("seal_code", "seal_name"), ("S-200", "Cartridge Seal"))},
        )

    monkeypatch.setattr(ExcelReader, "_read_xls", staticmethod(fake_read_xls))

    result = parse_import_file(path)

    assert isinstance(result, ImportPackage)
    assert result.seals == ({"seal_code": "S-200", "seal_name": "Cartridge Seal"},)


# --- Pump / Seal / Installation / Document mapping --------------------------------


def test_pump_sheet_maps_canonical_fields_case_insensitively(tmp_path):
    path = tmp_path / "pumps.xlsx"
    _save_workbook(
        path,
        [("Pump", (("Tag Number", "AREA", "Pump Type"), ("P-100", "Unit 1", "Centrifugal")))],
    )

    result = parse_import_file(path)

    assert result.pumps == ({"tag_number": "P-100", "area": "Unit 1", "pump_type": "Centrifugal"},)


def test_master_pump_sheet_name_is_recognized_as_the_pumps_entity(tmp_path):
    # MWO-LTSA-DATA-IMPORT-UI-001A: the real RU II acceptance workbook
    # (LTSA_RU_II_Master_Pump_Canonical.xlsx) names its data sheet "Master
    # Pump", not "Pump"/"Pumps" -- confirmed by direct inspection of that
    # real file. Same canonical "pumps" bucket, no new mapping vocabulary.
    path = tmp_path / "master_pump.xlsx"
    _save_workbook(
        path,
        [("Master Pump", (("Tag Number", "Area"), ("101-P-10A", "HSC")))],
    )

    result = parse_import_file(path)

    assert result.pumps == ({"tag_number": "101-P-10A", "area": "HSC"},)


def test_seal_sheet_maps_canonical_fields(tmp_path):
    path = tmp_path / "seals.xlsx"
    _save_workbook(
        path,
        [("Seals", (("seal_code", "seal_name", "manufacturer"), ("S-100", "Cartridge Seal", "John Crane")))],
    )

    result = parse_import_file(path)

    assert result.seals == ({"seal_code": "S-100", "seal_name": "Cartridge Seal", "manufacturer": "John Crane"},)


def test_installation_sheet_maps_canonical_fields(tmp_path):
    path = tmp_path / "installations.xlsx"
    _save_workbook(
        path,
        [
            (
                "Installation",
                (
                    ("installation_code", "report_no", "source_document_name"),
                    ("INSTL-001-2026", "001/INSTL/2026", "Install Report.pdf"),
                ),
            )
        ],
    )

    result = parse_import_file(path)

    assert result.installations == (
        {
            "installation_code": "INSTL-001-2026",
            "report_no": "001/INSTL/2026",
            "source_document_name": "Install Report.pdf",
        },
    )


def test_document_sheet_maps_canonical_fields(tmp_path):
    path = tmp_path / "documents.xlsx"
    _save_workbook(
        path,
        [
            (
                "Documents",
                (
                    ("document_code", "seal_code", "document_type", "title"),
                    ("DOC-100", "S-100", "DATASHEET", "Seal Datasheet"),
                ),
            )
        ],
    )

    result = parse_import_file(path)

    assert result.documents == (
        {"document_code": "DOC-100", "seal_code": "S-100", "document_type": "DATASHEET", "title": "Seal Datasheet"},
    )


# --- Multiple sheets / unknown sheet / unknown columns -----------------------------


def test_multiple_sheets_populate_multiple_entity_buckets(tmp_path):
    path = tmp_path / "multi.xlsx"
    _save_workbook(
        path,
        [
            ("Pump", (("tag_number", "area"), ("P-100", "Unit 1"))),
            ("Seal", (("seal_code", "seal_name"), ("S-100", "Cartridge Seal"))),
        ],
    )

    result = parse_import_file(path)

    assert result.pumps == ({"tag_number": "P-100", "area": "Unit 1"},)
    assert result.seals == ({"seal_code": "S-100", "seal_name": "Cartridge Seal"},)
    assert result.installations == ()
    assert result.documents == ()


def test_sheet_name_matching_is_case_insensitive(tmp_path):
    path = tmp_path / "case.xlsx"
    _save_workbook(path, [("PUMPS", (("tag_number", "area"), ("P-100", "Unit 1")))])

    result = parse_import_file(path)

    assert result.pumps == ({"tag_number": "P-100", "area": "Unit 1"},)


def test_unknown_sheet_is_ignored(tmp_path):
    path = tmp_path / "unknown-sheet.xlsx"
    _save_workbook(path, [("RandomSheet", (("tag_number", "area"), ("P-100", "Unit 1")))])

    result = parse_import_file(path)

    assert result == _EMPTY_PACKAGE


def test_unknown_columns_are_ignored(tmp_path):
    path = tmp_path / "unknown-columns.xlsx"
    _save_workbook(
        path,
        [("Pump", (("tag_number", "area", "Random Note"), ("P-100", "Unit 1", "irrelevant")))],
    )

    result = parse_import_file(path)

    assert result.pumps == ({"tag_number": "P-100", "area": "Unit 1"},)


# --- Empty workbook ----------------------------------------------------------------


def test_empty_workbook_produces_an_empty_import_package(tmp_path):
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)  # default single "Sheet" worksheet, unmatched name, zero rows

    result = parse_import_file(path)

    assert result == _EMPTY_PACKAGE


def test_a_matched_sheet_with_only_a_header_row_produces_no_records(tmp_path):
    path = tmp_path / "header-only.xlsx"
    _save_workbook(path, [("Pump", (("tag_number", "area"),))])

    result = parse_import_file(path)

    assert result.pumps == ()


# --- Corrupt workbook ----------------------------------------------------------------


def test_corrupt_xlsx_raises_manufacturing_validation_error(tmp_path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"this is not a real zip/xlsx file")

    with pytest.raises(ManufacturingValidationError):
        parse_import_file(path)


def test_corrupt_xls_raises_manufacturing_validation_error(tmp_path):
    path = tmp_path / "corrupt.xls"
    path.write_bytes(b"this is not a real BIFF/OLE2 file")

    with pytest.raises(ManufacturingValidationError):
        parse_import_file(path)


# --- Determinism ---------------------------------------------------------------------


def test_excel_adapter_is_deterministic_across_repeated_calls(tmp_path):
    path = tmp_path / "deterministic.xlsx"
    _save_workbook(path, [("Pump", (("tag_number", "area"), ("P-100", "Unit 1")))])

    first = parse_import_file(path)
    second = parse_import_file(path)

    assert first == second
