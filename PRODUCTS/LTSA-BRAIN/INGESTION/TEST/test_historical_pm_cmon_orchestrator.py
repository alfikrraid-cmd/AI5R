"""MWO-LTSA-HISTORICAL-JULY-INGESTION-001 -- end-to-end orchestrator
coverage against a SYNTHETIC workbook built in-memory with openpyxl,
shaped exactly like the real July "CM Measuring Report"/" PM Mech Seal"/
"Findings" sheets (per Phase 23's explicit synthetic-fixture allowance --
never a real customer PDF/XLSX in the test suite). Proves the full
register -> project -> match -> classify -> stage-candidate wiring is
correct in one pass, independent of any real source file's presence on
this machine.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from historical_pm_cmon_orchestrator import (  # noqa: E402
    build_area_dry_run,
    register_source_document,
)

_CANONICAL_TAGS = {"110-P-9A", "110-P-9B", "140-P-16A"}


def _build_synthetic_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cm = wb.create_sheet("CM Measuring Report")
    # Real layout: data starts row 11. Columns per _CM_COLUMNS.
    cm.cell(row=11, column=2, value=46570)  # date, an Excel serial (~2027-07-08, any parseable serial)
    cm.cell(row=11, column=3, value="110-P-9A")
    cm.cell(row=11, column=4, value="32/61")
    cm.cell(row=11, column=17, value="58")  # mechseal_temp_de
    cm.cell(row=11, column=18, value="58")  # mechseal_temp_nde
    cm.cell(row=11, column=19, value="Y")  # mechanical_seal_leak_de
    cm.cell(row=11, column=20, value="N")  # mechanical_seal_leak_nde

    cm.cell(row=12, column=2, value=46570)
    cm.cell(row=12, column=3, value="999-P-99Z")  # deliberately NOT in canonical roster
    cm.cell(row=12, column=4, value="53A")

    pm = wb.create_sheet(" PM Mech Seal")  # leading space, verbatim real sheet name
    pm.cell(row=7, column=9, value="1")  # checklist item-code header row
    pm.cell(row=11, column=2, value=46570)
    pm.cell(row=11, column=3, value="140-P-16A")
    pm.cell(row=11, column=4, value="MA-1")
    pm.cell(row=11, column=5, value="HOC")
    pm.cell(row=11, column=7, value="32/61")
    pm.cell(row=11, column=8, value="DONE")
    pm.cell(row=11, column=9, value="1")  # checklist item "1" done

    findings = wb.create_sheet("Findings")
    findings.cell(row=8, column=1, value="NO")
    findings.cell(row=8, column=2, value="AREA")
    findings.cell(row=8, column=3, value="TAG NO.")
    findings.cell(row=8, column=4, value="API PLAN")
    findings.cell(row=8, column=5, value="PUMP TYPE")
    findings.cell(row=8, column=6, value="SEAL LEAKAGE")
    findings.cell(row=8, column=12, value="REMARKS")
    findings.cell(row=8, column=13, value="N ")
    findings.cell(row=9, column=6, value="DE")
    findings.cell(row=9, column=7, value="NDE")
    findings.cell(row=11, column=1, value="1")
    findings.cell(row=11, column=2, value="DCU")
    findings.cell(row=11, column=3, value="140-P-16A")
    findings.cell(row=11, column=4, value="32/61")
    findings.cell(row=11, column=5, value="OH")
    findings.cell(row=11, column=6, value="Y")
    findings.cell(row=11, column=7, value="-")
    findings.cell(row=11, column=12, value="STANDBY, bocor dari draingland 1/2 detik")
    findings.cell(row=11, column=13, value="-")

    wb.save(path)


@pytest.fixture()
def synthetic_source(tmp_path) -> Path:
    xlsx_path = tmp_path / "Laporan PM, CM & Pemasangan Seal SYNTH JULY 2026.xlsx"
    pdf_path = tmp_path / "Laporan PM, CM & Pemasangan Seal SYNTH JULY 2026.pdf"
    _build_synthetic_workbook(xlsx_path)
    pdf_path.write_bytes(b"%PDF-1.4 synthetic placeholder, never parsed by the dry run\n")
    return register_source_document(pdf_path, xlsx_path, area_label="SYNTH")


class TestBuildAreaDryRun:
    def test_produces_expected_candidate_counts(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        assert len(result.pm_candidates) == 1
        assert len(result.cmon_candidates) == 2
        assert len(result.finding_candidates) == 1

    def test_document_classification_is_mixed(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        assert result.document_classification == "MIXED"

    def test_matched_tag_is_exact_match(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        matched = next(c for c in result.cmon_candidates if c.tag_number == "110-P-9A")
        assert matched.pump_match == "EXACT_MATCH"

    def test_unmatched_tag_is_no_match_never_silently_dropped(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        unmatched = next(c for c in result.cmon_candidates if c.tag_number == "999-P-99Z")
        assert unmatched.pump_match == "NO_MATCH"
        assert any("999-P-99Z" in item for item in result.critical_missing)

    def test_finding_candidate_carries_matched_tag_on_exact_match(self, synthetic_source):
        # MWO-LTSA-HISTORICAL-JULY-PROMOTION-001 rehearsal regression: the
        # Finding branch never passed matched_tag=match.outcome's own
        # resolved canonical tag into CandidateSummary (unlike the PM and
        # CMON branches), so stage_area_candidates() -- which requires
        # candidate.matched_tag to set pump_tag_number for EXACT_MATCH --
        # silently staged every EXACT_MATCH finding as pump_tag_number
        # NULL (wrongly classified INCOMPLETE). Real July data caught this:
        # 23/23 findings were EXACT_MATCH yet all 23 landed INCOMPLETE.
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        finding = result.finding_candidates[0]
        assert finding.pump_match == "EXACT_MATCH"
        assert finding.matched_tag == "140-P-16A"

    def test_finding_candidate_carries_real_remarks_never_the_wrong_column(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        finding = result.finding_candidates[0]
        assert finding.fields["remarks"] == "STANDBY, bocor dari draingland 1/2 detik"
        assert finding.fields["remarks"] != "32/61"  # the real defect this guards against (API PLAN column leak)

    def test_cmon_candidate_never_carries_a_finding_field_at_extraction_time(self, synthetic_source):
        # Findings are staged as their own candidates, never auto-merged
        # onto a CMON reading (no reliable per-row date to disambiguate
        # which of that tag's multiple readings it belongs to).
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        assert all("finding" not in c.fields for c in result.cmon_candidates)

    def test_blank_measurement_stays_none_never_fabricated(self, synthetic_source):
        result = build_area_dry_run(synthetic_source, canonical_pump_tags=_CANONICAL_TAGS)
        second_row = next(c for c in result.cmon_candidates if c.tag_number == "999-P-99Z")
        assert second_row.fields.get("mechseal_temp_de") is None

    def test_exact_duplicate_source_is_not_staged(self, synthetic_source):
        result = build_area_dry_run(
            synthetic_source, canonical_pump_tags=_CANONICAL_TAGS,
            known_source_hashes={synthetic_source.xlsx_sha256},
        )
        assert result.pm_candidates == []
        assert result.cmon_candidates == []
        assert result.finding_candidates == []
        assert any("EXACT_DUPLICATE" in item for item in result.critical_missing)


class TestRegisterSourceDocument:
    def test_hashes_are_real_and_stable(self, synthetic_source):
        assert len(synthetic_source.pdf_sha256) == 64
        assert len(synthetic_source.xlsx_sha256) == 64
