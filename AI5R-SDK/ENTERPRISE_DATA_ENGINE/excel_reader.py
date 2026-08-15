from pathlib import Path

import xlrd
from openpyxl import load_workbook

from .import_error import ImportError
from .raw_dataset import RawDataset, ReaderMetadata
from .reader import Reader
from .source_descriptor import SourceDescriptor


class ExcelReader(Reader):
    """Acquire raw worksheet values from .xlsx and .xls workbooks."""

    __slots__ = ()

    _MEDIA_TYPES = {
        ".xlsx": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        ".xls": "application/vnd.ms-excel",
    }

    def supports(self, source: SourceDescriptor) -> bool:
        return Path(source.location).suffix.lower() in self._MEDIA_TYPES

    def read(self, source: SourceDescriptor) -> RawDataset:
        suffix = Path(source.location).suffix.lower()
        if suffix not in self._MEDIA_TYPES:
            raise ImportError(
                code="UNSUPPORTED_SOURCE",
                message=f"ExcelReader does not support: {source.location}",
            )

        if suffix == ".xlsx":
            worksheets = self._read_xlsx(Path(source.location))
        else:
            worksheets = self._read_xls(Path(source.location))

        worksheet_names = tuple(sheet["name"] for sheet in worksheets)
        return RawDataset(
            data=worksheets,
            metadata=ReaderMetadata(
                reader_name="excel",
                source_type=suffix.removeprefix("."),
                media_type=self._MEDIA_TYPES[suffix],
                attributes={
                    "worksheet_count": len(worksheets),
                    "worksheet_names": worksheet_names,
                },
            ),
        )

    @staticmethod
    def _read_xlsx(path: Path) -> tuple[dict, ...]:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
        )
        try:
            return tuple(
                {
                    "name": worksheet.title,
                    "rows": tuple(
                        tuple(row)
                        for row in worksheet.iter_rows(values_only=True)
                    ),
                }
                for worksheet in workbook.worksheets
            )
        finally:
            workbook.close()

    @staticmethod
    def _read_xls(path: Path) -> tuple[dict, ...]:
        workbook = xlrd.open_workbook(path)
        return tuple(
            {
                "name": worksheet.name,
                "rows": tuple(
                    tuple(
                        worksheet.cell_value(row_index, column_index)
                        for column_index in range(worksheet.ncols)
                    )
                    for row_index in range(worksheet.nrows)
                ),
            }
            for worksheet in workbook.sheets()
        )


__all__ = ["ExcelReader"]
